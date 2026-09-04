"""Genera nodos, elementos, Excel y CSV desde geometria_manual.json."""

from pathlib import Path
import json

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUTPUTS = ROOT / "outputs"
MODEL = OUTPUTS / "modelo_3d_manual.json"
EXCEL = OUTPUTS / "nodos_modelo_manual.xlsx"
CSV = ROOT / "UnityVisualization" / "Assets" / "Resources" / "model_3d.csv"
LOADS = DATA / "cargas_losas.json"
KG_TO_KN = 9.80665 / 1000.0


def main():
    config = json.loads((DATA / "geometria_manual.json").read_text(encoding="utf-8"))
    loads = json.loads(LOADS.read_text(encoding="utf-8")) if LOADS.exists() else {"levels": [], "lt2": {}, "lt2_load_cases": []}
    slab_thickness = config.get("slab_thickness_m", 0.15)
    slab_density = config.get("slab_density_kg_m3", 2500.0)
    slab_self_weight = slab_thickness * slab_density * 9.80665 / 1000.0
    nodes = []
    by_level = {z: [] for z in config["levels_m"]}
    for k, z in enumerate(config["levels_m"]):
        for y_index, y in enumerate(config["y_axes_m"]):
            for axis in config["column_axes"]:
                if z < axis["z_inicio_m"]:
                    continue
                tag = 100000 * k + 1000 * y_index + int(axis["x_m"] * 10) + 1
                node = {"id": tag, "level": k, "axis": axis["eje"],
                        "x_m": axis["x_m"], "y_m": y, "z_m": z,
                        "restraint": k == 0, "status": "MANUAL"}
                nodes.append(node)
                by_level[z].append(node)

        # Some columns occupy only selected intersections of the extended grid.
        # They are declared explicitly instead of being created at every Y row.
        for point_index, point in enumerate(config.get("column_points", [])):
            if z < point.get("z_inicio_m", 0.0):
                continue
            tag = 700000 + 1000 * k + point_index + 1
            node = {"id": tag, "level": k, "axis": point["id"],
                    "x_m": point["x_m"], "y_m": point["y_m"], "z_m": z,
                    "restraint": k == 0, "status": "MANUAL"}
            nodes.append(node)
            by_level[z].append(node)

    wall_nodes = {}
    wall_node_base = 600000
    for wall_index, wall in enumerate(config.get("walls", [])):
        z0, z1 = wall["z_i_m"], wall["z_j_m"]
        if z0 == z1:
            continue
        for level_index, z in enumerate(config["levels_m"]):
            if not min(z0, z1) <= z <= max(z0, z1):
                continue
            ratio = (z - z0) / (z1 - z0)
            x = wall["x_i_m"] + ratio * (wall["x_j_m"] - wall["x_i_m"])
            y = wall["y_i_m"] + ratio * (wall["y_j_m"] - wall["y_i_m"])
            tag = wall_node_base + wall_index * 1000 + level_index + 1
            node = {"id": tag, "level": level_index, "axis": f"W{wall['id']}",
                    "x_m": x, "y_m": y, "z_m": z, "restraint": level_index == 0,
                    "status": "MANUAL_WALL_NODE"}
            nodes.append(node)
            by_level[z].append(node)
            wall_nodes[(wall_index, round(z, 3))] = node

    manual_beams = list(config.get("manual_beams", []))
    for repeat in config.get("repeat_manual_beams", []):
        def allowed(beam):
            if not repeat.get("exclude_negative_y", False):
                return True
            return all(beam.get(key, 0.0) >= 0.0 for key in ("y_m", "y_i_m", "y_j_m") if key in beam)

        manual_beams.extend(
            [dict(beam, z_m=repeat["to_z_m"])
             for beam in manual_beams
             if round(beam["z_m"], 3) == round(repeat["from_z_m"], 3)
             and (not repeat.get("source_group") or beam.get("group") == repeat["source_group"])
             and beam.get("group") != repeat.get("exclude_group")
             and allowed(beam)]
        )

    elements = []
    eid = 1
    levels = config["levels_m"]
    for z0, z1 in zip(levels, levels[1:]):
        low = {(n["axis"], round(n["y_m"], 3)): n for n in by_level[z0]
               if n["status"] != "MANUAL_WALL_NODE"}
        high = {(n["axis"], round(n["y_m"], 3)): n for n in by_level[z1]
                if n["status"] != "MANUAL_WALL_NODE"}
        for key, top in high.items():
            if key in low:
                elements.append({"id": eid, "type": "COLUMN", "i": low[key]["id"], "j": top["id"], "status": "MANUAL"})
                eid += 1

    wall_edge_ids = {}
    for wall_index, wall in enumerate(config.get("walls", [])):
        for z0, z1 in zip(levels, levels[1:]):
            a = wall_nodes.get((wall_index, round(z0, 3)))
            b = wall_nodes.get((wall_index, round(z1, 3)))
            if a and b:
                elements.append({"id": eid, "type": "WALL", "i": a["id"], "j": b["id"],
                                 "thickness_m": wall["thickness_m"], "status": "MANUAL_WALL"})
                wall_edge_ids[(wall_index, round(z0, 3))] = eid
                eid += 1

    manual_z = {round(b["z_m"], 3) for b in manual_beams}
    node_at = {(n["axis"], round(n["y_m"], 3), round(n["z_m"], 3)): n for n in nodes}
    node_by_xyz = {(round(n["x_m"], 3), round(n["y_m"], 3), round(n["z_m"], 3)): n for n in nodes}

    # Connect wall line nodes to nearby frame nodes when the wall thickness
    # produces a small centerline offset in the source geometry.
    frame_nodes = [node for node in nodes if node["status"] != "MANUAL_WALL_NODE"]
    for element in elements:
        if element["type"] != "WALL":
            continue
        for endpoint in ("i", "j"):
            wall_node = next(node for node in nodes if node["id"] == element[endpoint])
            candidates = [node for node in frame_nodes + [candidate for candidate in nodes
                                                          if candidate["status"] == "MANUAL_WALL_NODE"]
                          if node["id"] != wall_node["id"]
                          if abs(node["z_m"] - wall_node["z_m"]) < 1e-6
                          and ((node["x_m"] - wall_node["x_m"]) ** 2
                               + (node["y_m"] - wall_node["y_m"]) ** 2) ** 0.5 <= 0.60]
            if candidates:
                nearest = min(candidates, key=lambda node: (
                    (node["x_m"] - wall_node["x_m"]) ** 2
                    + (node["y_m"] - wall_node["y_m"]) ** 2,
                    node["id"],
                ))
                element[endpoint] = nearest["id"]

    next_manual_node = 900000

    def ensure_beam_node(x, y, z):
        nonlocal next_manual_node
        key = (round(x, 3), round(y, 3), round(z, 3))
        if key not in node_by_xyz:
            node = {"id": next_manual_node, "level": levels.index(z), "axis": "BEAM_MANUAL",
                    "x_m": x, "y_m": y, "z_m": z, "restraint": False, "status": "MANUAL_BEAM_NODE"}
            next_manual_node += 1
            nodes.append(node)
            node_by_xyz[key] = node
        return node_by_xyz[key]

    for beam in manual_beams:
        z = round(beam["z_m"], 3)
        if "x_i_m" in beam:
            a = ensure_beam_node(beam["x_i_m"], beam["y_i_m"], z)
            b = ensure_beam_node(beam["x_j_m"], beam["y_j_m"], z)
        elif beam["direction"] == "X":
            a = node_at[(beam["axis_i"], round(beam["y_m"], 3), z)]
            b = node_at[(beam["axis_j"], round(beam["y_m"], 3), z)]
        else:
            x = next(axis["x_m"] for axis in config["column_axes"] if axis["eje"] == beam["axis"])
            a = node_at.get((beam["axis"], round(beam["y_i_m"], 3), z),
                            ensure_beam_node(x, beam["y_i_m"], z))
            b = node_at.get((beam["axis"], round(beam["y_j_m"], 3), z),
                            ensure_beam_node(x, beam["y_j_m"], z))
        elements.append({"id": eid, "type": beam.get("member_type", "BEAM_" + beam["direction"]), "direction": beam["direction"], "i": a["id"], "j": b["id"], "status": "MANUAL"})
        eid += 1

    # Provisional beams: connect adjacent listed column axes and Y rows where no manual layout exists.
    for z in levels[1:]:
        if round(z, 3) in manual_z:
            continue
        current = by_level[z]
        for y in config["y_axes_m"]:
            row = sorted((n for n in current if abs(n["y_m"] - y) < 1e-9), key=lambda n: n["x_m"])
            for a, b in zip(row, row[1:]):
                elements.append({"id": eid, "type": "BEAM_X", "direction": "X", "i": a["id"], "j": b["id"], "status": "PROVISIONAL_BEAM"})
                eid += 1
        for axis in config["column_axes"]:
            col = sorted((n for n in current if n["axis"] == axis["eje"]), key=lambda n: n["y_m"])
            for a, b in zip(col, col[1:]):
                elements.append({"id": eid, "type": "BEAM_Y", "direction": "Y", "i": a["id"], "j": b["id"], "status": "PROVISIONAL_BEAM"})
                eid += 1

    # Split coplanar beam crossings so every physical intersection has a shared node.
    beam_elements = [element for element in elements if element["type"].startswith("BEAM")]
    split_elements = []
    next_split_element = eid

    def beam_coordinates(element):
        a = next(n for n in nodes if n["id"] == element["i"])
        b = next(n for n in nodes if n["id"] == element["j"])
        return a, b

    for element in elements:
        if not element["type"].startswith("BEAM"):
            split_elements.append(element)
            continue
        a, b = beam_coordinates(element)
        cuts = [(0.0, a["x_m"], a["y_m"]), (1.0, b["x_m"], b["y_m"])]
        horizontal = abs(a["y_m"] - b["y_m"]) < 1e-6
        for other in beam_elements:
            if other["id"] == element["id"]:
                continue
            c, d = beam_coordinates(other)
            if abs(c["z_m"] - a["z_m"]) > 1e-6 or abs(d["z_m"] - a["z_m"]) > 1e-6:
                continue
            other_horizontal = abs(c["y_m"] - d["y_m"]) < 1e-6
            if horizontal and not other_horizontal and min(a["x_m"], b["x_m"]) - 1e-6 <= c["x_m"] <= max(a["x_m"], b["x_m"]) + 1e-6 and min(c["y_m"], d["y_m"]) - 1e-6 <= a["y_m"] <= max(c["y_m"], d["y_m"]) + 1e-6:
                ratio = (c["x_m"] - a["x_m"]) / (b["x_m"] - a["x_m"])
                cuts.append((ratio, c["x_m"], a["y_m"]))
            elif not horizontal and other_horizontal and min(c["x_m"], d["x_m"]) - 1e-6 <= a["x_m"] <= max(c["x_m"], d["x_m"]) + 1e-6 and min(a["y_m"], b["y_m"]) - 1e-6 <= c["y_m"] <= max(a["y_m"], b["y_m"]) + 1e-6:
                ratio = (c["y_m"] - a["y_m"]) / (b["y_m"] - a["y_m"])
                cuts.append((ratio, a["x_m"], c["y_m"]))
        for wall_node in nodes:
            if wall_node["status"] != "MANUAL_WALL_NODE" or abs(wall_node["z_m"] - a["z_m"]) > 1e-6:
                continue
            if horizontal and abs(wall_node["y_m"] - a["y_m"]) < 1e-6 and min(a["x_m"], b["x_m"]) - 1e-6 <= wall_node["x_m"] <= max(a["x_m"], b["x_m"]) + 1e-6:
                ratio = (wall_node["x_m"] - a["x_m"]) / (b["x_m"] - a["x_m"])
                cuts.append((ratio, wall_node["x_m"], wall_node["y_m"]))
            elif not horizontal and abs(wall_node["x_m"] - a["x_m"]) < 1e-6 and min(a["y_m"], b["y_m"]) - 1e-6 <= wall_node["y_m"] <= max(a["y_m"], b["y_m"]) + 1e-6:
                ratio = (wall_node["y_m"] - a["y_m"]) / (b["y_m"] - a["y_m"])
                cuts.append((ratio, wall_node["x_m"], wall_node["y_m"]))
        unique_cuts = sorted({(round(r, 9), round(x, 6), round(y, 6)) for r, x, y in cuts})
        cut_nodes = [ensure_beam_node(x, y, a["z_m"]) for _, x, y in unique_cuts]
        for index, (node_a, node_b) in enumerate(zip(cut_nodes, cut_nodes[1:])):
            if node_a["id"] == node_b["id"]:
                continue
            split_elements.append({**element, "id": element["id"] if index == 0 else next_split_element,
                                   "i": node_a["id"], "j": node_b["id"],
                                   "status": "SPLIT_INTERSECTION" if len(cut_nodes) > 2 else element["status"]})
            if index > 0:
                next_split_element += 1
    elements = split_elements
    eid = next_split_element

    def covered(intervals, start, end):
        current = start
        for a, b in sorted((min(a, b), max(a, b)) for a, b in intervals):
            if b <= current + 1e-6:
                continue
            if a > current + 1e-6:
                return False
            current = max(current, b)
            if current >= end - 1e-6:
                return True
        return current >= end - 1e-6

    def covered_with_ids(intervals, start, end):
        current = start
        ids = []
        for a, b, element_id in sorted((min(a, b), max(a, b), element_id) for a, b, element_id in intervals):
            if b <= current + 1e-6:
                continue
            if a > current + 1e-6:
                return []
            ids.append(element_id)
            current = max(current, b)
            if current >= end - 1e-6:
                return ids
        return ids if current >= end - 1e-6 else []

    def tributary_loads_for_rectangle(lx, ly, edges):
        area = lx * ly
        loads = []

        def add(edge, tributary_area, distribution, w_start, w_max, w_end):
            loads.append({
                "edge": edge,
                "beam_ids": edges[edge],
                "tributary_area_m2": round(tributary_area, 6),
                "distribution": distribution,
                "w_start_kN_m": round(w_start, 6),
                "w_max_kN_m": round(w_max, 6),
                "w_end_kN_m": round(w_end, 6),
                "total_load_kN": round(slab_self_weight * tributary_area, 6),
            })

        if lx >= ly:
            short = ly
            long = lx
            short_edge_area = short * short / 4.0
            long_edge_area = short * (2.0 * long - short) / 4.0
            w_max = slab_self_weight * short / 2.0
            add("left", short_edge_area, "triangular", 0.0, w_max, 0.0)
            add("right", short_edge_area, "triangular", 0.0, w_max, 0.0)
            add("bottom", long_edge_area, "trapezoidal", 0.0, w_max, 0.0)
            add("top", long_edge_area, "trapezoidal", 0.0, w_max, 0.0)
        else:
            short = lx
            long = ly
            short_edge_area = short * short / 4.0
            long_edge_area = short * (2.0 * long - short) / 4.0
            w_max = slab_self_weight * short / 2.0
            add("bottom", short_edge_area, "triangular", 0.0, w_max, 0.0)
            add("top", short_edge_area, "triangular", 0.0, w_max, 0.0)
            add("left", long_edge_area, "trapezoidal", 0.0, w_max, 0.0)
            add("right", long_edge_area, "trapezoidal", 0.0, w_max, 0.0)

        return loads

    # Detect every elementary rectangular face bounded by four coplanar beams.
    slab_faces = []
    for z in levels[1:]:
        horizontal = {}
        vertical = {}
        x_breaks = set()
        y_breaks = set()
        for element in elements:
            if not element["type"].startswith("BEAM"):
                continue
            a = next(n for n in nodes if n["id"] == element["i"])
            b = next(n for n in nodes if n["id"] == element["j"])
            if abs(a["z_m"] - z) > 1e-6 or abs(b["z_m"] - z) > 1e-6:
                continue
            x_breaks.update((round(a["x_m"], 3), round(b["x_m"], 3)))
            y_breaks.update((round(a["y_m"], 3), round(b["y_m"], 3)))
            if abs(a["y_m"] - b["y_m"]) < 1e-6:
                horizontal.setdefault(round(a["y_m"], 6), []).append((a["x_m"], b["x_m"], element["id"]))
            elif abs(a["x_m"] - b["x_m"]) < 1e-6:
                vertical.setdefault(round(a["x_m"], 6), []).append((a["y_m"], b["y_m"], element["id"]))
        for wall_index, wall in enumerate(config.get("walls", [])):
            if not min(wall["z_i_m"], wall["z_j_m"]) - 1e-6 <= z <= max(wall["z_i_m"], wall["z_j_m"]) + 1e-6:
                wall_id = wall_edge_ids.get((wall_index, round(max(
                    level_value for level_value in levels if level_value <= z + 1e-6
                ), 3)))
                if wall_id is None:
                    continue
                if abs(wall["y_i_m"] - wall["y_j_m"]) < 1e-6:
                    horizontal.setdefault(round(wall["y_i_m"], 6), []).append((wall["x_i_m"], wall["x_j_m"], wall_id))
                    x_breaks.update((round(wall["x_i_m"], 3), round(wall["x_j_m"], 3)))
                    y_breaks.add(round(wall["y_i_m"], 3))
                elif abs(wall["x_i_m"] - wall["x_j_m"]) < 1e-6:
                    vertical.setdefault(round(wall["x_i_m"], 6), []).append((wall["y_i_m"], wall["y_j_m"], wall_id))
                    x_breaks.add(round(wall["x_i_m"], 3))
                    y_breaks.update((round(wall["y_i_m"], 3), round(wall["y_j_m"], 3)))
        xs = sorted(x_breaks)
        ys = sorted(y_breaks)
        for i, x0 in enumerate(xs):
            for x1 in xs[i + 1:]:
                if x1 - x0 < 0.05:
                    continue
                for j, y0 in enumerate(ys):
                    for y1 in ys[j + 1:]:
                        if y1 - y0 < 0.05:
                            continue
                        edges = {
                            "bottom": covered_with_ids(horizontal.get(round(y0, 6), []), x0, x1),
                            "top": covered_with_ids(horizontal.get(round(y1, 6), []), x0, x1),
                            "left": covered_with_ids(vertical.get(round(x0, 6), []), y0, y1),
                            "right": covered_with_ids(vertical.get(round(x1, 6), []), y0, y1),
                        }
                        if not all(edges.values()):
                            continue
                        has_internal_horizontal = any(
                            y0 + 1e-6 < y < y1 - 1e-6
                            and covered_with_ids(intervals, x0, x1)
                            for y, intervals in horizontal.items()
                        )
                        has_internal_vertical = any(
                            x0 + 1e-6 < x < x1 - 1e-6
                            and covered_with_ids(intervals, y0, y1)
                            for x, intervals in vertical.items()
                        )
                        if has_internal_horizontal or has_internal_vertical:
                            continue
                        slab_faces.append({
                            "z_m": z,
                            "vertices": [(x0, y0), (x1, y0), (x1, y1), (x0, y1)],
                            "edge_beam_ids": edges,
                        })

    slabs = []
    next_slab_node = 800000
    next_slab_element = 1000000
    for face_data in slab_faces:
        z = face_data["z_m"]
        face = face_data["vertices"]
        slab_node_ids = []
        for x, y in face:
            key = (round(x, 3), round(y, 3), round(z, 3))
            if key not in node_by_xyz:
                node = {"id": next_slab_node, "level": levels.index(z), "axis": "SLAB",
                        "x_m": x, "y_m": y, "z_m": z, "restraint": False,
                        "status": "SLAB_NODE"}
                next_slab_node += 1
                nodes.append(node)
                node_by_xyz[key] = node
            slab_node_ids.append(node_by_xyz[key]["id"])
        lx = abs(face[1][0] - face[0][0])
        ly = abs(face[2][1] - face[1][1])
        tributary_loads = tributary_loads_for_rectangle(lx, ly, face_data["edge_beam_ids"])
        area = lx * ly
        assigned_area = sum(load["tributary_area_m2"] for load in tributary_loads)
        boundary_elements = sorted({element_id for edge_ids in face_data["edge_beam_ids"].values()
                                    for element_id in edge_ids})
        slabs.append({"id": next_slab_element, "panel_id": next_slab_element,
                       "node_ids": slab_node_ids,
                      "coordinates": [{"x_m": x, "y_m": y, "z_m": z} for x, y in face],
                      "edge_beam_ids": face_data["edge_beam_ids"],
                      "dimensions_m": {"lx": round(lx, 6), "ly": round(ly, 6)},
                      "area_m2": round(area, 6),
                       "thickness_m": slab_thickness,
                       "density_kg_m3": slab_density,
                       "self_weight_kN_m2": round(slab_self_weight, 6),
                      "tributary_loads": tributary_loads,
                      "area_check_m2": round(assigned_area - area, 6),
                      "boundary_elements": boundary_elements,
                      "z_m": z, "status": "AUTO_CLOSED_BY_BEAMS_AND_WALLS"})
        next_slab_element += 1

    def point_in_polygon_xy(point, polygon):
        x, y = point
        inside = False
        for i, (x1, y1) in enumerate(polygon):
            x2, y2 = polygon[(i + 1) % len(polygon)]
            if (y1 > y) != (y2 > y):
                crossing_x = (x2 - x1) * (y - y1) / (y2 - y1) + x1
                if x < crossing_x:
                    inside = not inside
        return inside

    def void_rectangles_for_slab(slab):
        x0 = min(point["x_m"] for point in slab["coordinates"])
        x1 = max(point["x_m"] for point in slab["coordinates"])
        y0 = min(point["y_m"] for point in slab["coordinates"])
        y1 = max(point["y_m"] for point in slab["coordinates"])
        geometries = []
        level = next((item for item in loads.get("levels", [])
                      if abs(item["z_m"] - slab["z_m"]) < 1e-6), None)
        if level:
            geometries.extend(
                geometry for zone in level.get("zones", [])
                for geometry in zone.get("geometries", [])
                if geometry.get("type") == "void"
            )
        lt2_geometries = loads.get("lt2", {}).get("geometries", [])
        active_lt2 = lt2_geometries if abs(slab["z_m"] - 19.8) < 1e-6 else lt2_geometries[:6]
        geometries.extend(
            geometry for geometry in active_lt2
            if geometry.get("type") == "void"
        )
        rectangles = []
        seen = set()
        for geometry in geometries:
            points = geometry.get("points", [])
            if not points:
                continue
            vx0, vx1 = min(point[0] for point in points), max(point[0] for point in points)
            vy0, vy1 = min(point[1] for point in points), max(point[1] for point in points)
            ix0, ix1 = max(x0, vx0), min(x1, vx1)
            iy0, iy1 = max(y0, vy0), min(y1, vy1)
            if ix1 - ix0 <= 1e-6 or iy1 - iy0 <= 1e-6:
                continue
            key = tuple(round(value, 6) for value in (ix0, ix1, iy0, iy1))
            if key not in seen:
                seen.add(key)
                rectangles.append({"x_min_m": ix0, "x_max_m": ix1,
                                   "y_min_m": iy0, "y_max_m": iy1,
                                   "area_m2": (ix1 - ix0) * (iy1 - iy0)})
        return rectangles

    void_slabs = []
    valid_slabs = []
    for slab in slabs:
        gross_area = slab["area_m2"]
        voids = void_rectangles_for_slab(slab)
        void_area = sum(void["area_m2"] for void in voids)
        effective_area = max(gross_area - void_area, 0.0)
        slab["gross_area_m2"] = gross_area
        slab["voids"] = voids
        slab["area_m2"] = round(effective_area, 6)
        if effective_area <= 1e-6:
            void_slabs.append(slab)
            continue
        slab["area_check_m2"] = round(gross_area - void_area - effective_area, 6)
        valid_slabs.append(slab)
    slabs = valid_slabs

    beam_slab_loads = []
    element_by_id = {element["id"]: element for element in elements}

    def beam_share(beam_ids, beam_id):
        lengths = {}
        for candidate_id in beam_ids:
            candidate = element_by_id[candidate_id]
            a = next(n for n in nodes if n["id"] == candidate["i"])
            b = next(n for n in nodes if n["id"] == candidate["j"])
            lengths[candidate_id] = ((a["x_m"] - b["x_m"]) ** 2 + (a["y_m"] - b["y_m"]) ** 2) ** 0.5
        total = sum(lengths.values())
        return lengths[beam_id] / total if total else 1.0 / len(beam_ids)

    for slab in slabs:
        for load in slab["tributary_loads"]:
            for beam_id in load["beam_ids"]:
                share = beam_share(load["beam_ids"], beam_id)
                beam_slab_loads.append({
                    "slab_id": slab["id"],
                    "beam_id": beam_id,
                    "z_m": slab["z_m"],
                    "edge": load["edge"],
                    "tributary_area_m2": round(load["tributary_area_m2"] * share, 6),
                    "distribution": load["distribution"],
                    "w_start_kN_m": round(load["w_start_kN_m"] * share, 6),
                    "w_max_kN_m": round(load["w_max_kN_m"] * share, 6),
                    "w_end_kN_m": round(load["w_end_kN_m"] * share, 6),
                    "total_load_kN": round(load["total_load_kN"] * share, 6),
                })

    def polygon_area(points):
        return abs(sum(
            points[i][0] * points[(i + 1) % len(points)][1]
            - points[(i + 1) % len(points)][0] * points[i][1]
            for i in range(len(points))
        )) / 2.0 if len(points) >= 3 else 0.0

    load_zone_totals = []
    for level in loads.get("levels", []):
        for zone in level.get("zones", []):
            area = 0.0
            for geometry in zone.get("geometries", []):
                contribution = polygon_area(geometry.get("points", []))
                area += contribution if geometry.get("type") == "polygon" else -contribution
            area = max(area, 0.0)
            pm = zone.get("pm_adic_kg_m2") or 0.0
            sc = zone.get("sc_kg_m2") or 0.0
            q_pm = pm * KG_TO_KN
            q_sc = sc * KG_TO_KN
            load_zone_totals.append({
                "level_z_m": level["z_m"],
                "zone": zone["id"],
                "area_m2": round(area, 6),
                "q_pp_kN_m2": round(slab_self_weight, 6),
                "q_pm_adic_kN_m2": round(q_pm, 6),
                "q_G_kN_m2": round(slab_self_weight + q_pm, 6),
                "q_sc_kN_m2": round(q_sc, 6),
                "dead_load_kN": round(area * (slab_self_weight + q_pm), 6),
                "live_load_kN": round(area * q_sc, 6),
            })

    def point_in_polygon(point, polygon):
        x, y = point
        inside = False
        for i, (x1, y1) in enumerate(polygon):
            x2, y2 = polygon[(i + 1) % len(polygon)]
            if (y1 > y) != (y2 > y):
                crossing_x = (x2 - x1) * (y - y1) / (y2 - y1) + x1
                if x < crossing_x:
                    inside = not inside
        return inside

    def zone_contains(zone, point):
        inside_outer = any(
            geometry["type"] == "polygon"
            and point_in_polygon(point, geometry["points"])
            for geometry in zone.get("geometries", [])
        )
        inside_void = any(
            geometry["type"] == "void"
            and point_in_polygon(point, geometry["points"])
            for geometry in zone.get("geometries", [])
        )
        return inside_outer and not inside_void

    zones_by_level = {level["z_m"]: level.get("zones", []) for level in loads.get("levels", [])}
    lt2_geometries = loads.get("lt2", {}).get("geometries", [])

    def lt2_zone_for(point, z):
        active_geometries = lt2_geometries if abs(z - 19.8) < 1e-6 else lt2_geometries[:6]
        if any(
            geometry.get("type") == "void"
            and point_in_polygon(point, geometry["points"])
            for geometry in active_geometries
        ):
            return None
        if z == 19.8 and len(active_geometries) >= 7:
            outer = lt2_geometries[6]
            passed = any(point_in_polygon(point, geometry["points"]) for geometry in active_geometries[7:])
            if outer.get("type") == "polygon" and point_in_polygon(point, outer["points"]) and not passed:
                return {"id": "LT2 piso 4", "pm_adic_kg_m2": 200.0, "sc_kg_m2": 200.0}
            return None
        if len(lt2_geometries) < 6:
            return None
        if point_in_polygon(point, lt2_geometries[4]["points"]):
            return {"id": "LT2 D", "pm_adic_kg_m2": 260.0, "sc_kg_m2": 300.0}
        for index, name in ((0, "LT2 A"), (1, "LT2 B"), (3, "LT2 C"), (5, "LT2 E")):
            geometry = lt2_geometries[index]
            if geometry.get("type") == "polygon" and point_in_polygon(point, geometry["points"]):
                sc = 200.0 if name == "LT2 E" else 500.0
                return {"id": name, "pm_adic_kg_m2": 260.0, "sc_kg_m2": sc}
        return None

    def clip_polygon(polygon, axis, bound, keep_greater):
        clipped = []
        for index, current in enumerate(polygon):
            previous = polygon[index - 1]
            current_value = current[0] if axis == "x" else current[1]
            previous_value = previous[0] if axis == "x" else previous[1]
            current_inside = current_value >= bound - 1e-9 if keep_greater else current_value <= bound + 1e-9
            previous_inside = previous_value >= bound - 1e-9 if keep_greater else previous_value <= bound + 1e-9
            if current_inside != previous_inside:
                denominator = current_value - previous_value
                ratio = (bound - previous_value) / denominator if abs(denominator) > 1e-12 else 0.0
                clipped.append((
                    previous[0] + ratio * (current[0] - previous[0]),
                    previous[1] + ratio * (current[1] - previous[1]),
                ))
            if current_inside:
                clipped.append(current)
        return clipped

    def polygon_area_xy(polygon):
        return abs(sum(
            polygon[index][0] * polygon[(index + 1) % len(polygon)][1]
            - polygon[(index + 1) % len(polygon)][0] * polygon[index][1]
            for index in range(len(polygon))
        )) / 2.0 if len(polygon) >= 3 else 0.0

    def polygon_intersection_area(subject, clip):
        if not subject or not clip:
            return 0.0
        signed_clip_area = sum(
            clip[index][0] * clip[(index + 1) % len(clip)][1]
            - clip[(index + 1) % len(clip)][0] * clip[index][1]
            for index in range(len(clip))
        )
        orientation = 1.0 if signed_clip_area >= 0.0 else -1.0
        result = list(subject)
        for index, boundary_start in enumerate(clip):
            boundary_end = clip[(index + 1) % len(clip)]

            def cross(point):
                return ((boundary_end[0] - boundary_start[0]) * (point[1] - boundary_start[1])
                        - (boundary_end[1] - boundary_start[1]) * (point[0] - boundary_start[0]))

            clipped = []
            for point_index, current in enumerate(result):
                previous = result[point_index - 1]
                current_inside = orientation * cross(current) >= -1e-9
                previous_inside = orientation * cross(previous) >= -1e-9
                if current_inside != previous_inside:
                    current_value = cross(current)
                    previous_value = cross(previous)
                    ratio = previous_value / (previous_value - current_value)
                    clipped.append((
                        previous[0] + ratio * (current[0] - previous[0]),
                        previous[1] + ratio * (current[1] - previous[1]),
                    ))
                if current_inside:
                    clipped.append(current)
            result = clipped
            if not result:
                return 0.0
        return polygon_area_xy(result)

    def polygon_rectangle_intersection_polygon(polygon, xa, xb, ya, yb):
        clipped = clip_polygon(polygon, "x", xa, True)
        clipped = clip_polygon(clipped, "x", xb, False)
        clipped = clip_polygon(clipped, "y", ya, True)
        clipped = clip_polygon(clipped, "y", yb, False)
        return clipped

    def polygon_rectangle_intersection_area(polygon, xa, xb, ya, yb):
        return polygon_area_xy(polygon_rectangle_intersection_polygon(polygon, xa, xb, ya, yb))

    def tributary_polygons(slab):
        points = slab["coordinates"]
        x0 = min(point["x_m"] for point in points)
        x1 = max(point["x_m"] for point in points)
        y0 = min(point["y_m"] for point in points)
        y1 = max(point["y_m"] for point in points)
        xm, ym = (x0 + x1) / 2.0, (y0 + y1) / 2.0
        if x1 - x0 >= y1 - y0:
            depth = (y1 - y0) / 2.0
            return {
                "left": [(x0, y0), (x0, y1), (x0 + depth, ym)],
                "right": [(x1, y0), (x1 - depth, ym), (x1, y1)],
                "bottom": [(x0, y0), (x1, y0), (x1 - depth, ym), (x0 + depth, ym)],
                "top": [(x0, y1), (x0 + depth, ym), (x1 - depth, ym), (x1, y1)],
            }
        depth = (x1 - x0) / 2.0
        return {
            "bottom": [(x0, y0), (x1, y0), (xm, y0 + depth)],
            "top": [(x0, y1), (xm, y1 - depth), (x1, y1)],
            "left": [(x0, y0), (xm, y0 + depth), (xm, y1 - depth), (x0, y1)],
            "right": [(x1, y0), (x1, y1), (xm, y1 - depth), (xm, y0 + depth)],
        }

    def construction_lines_45(slab, polygons):
        points = slab["coordinates"]
        x0 = min(point["x_m"] for point in points)
        x1 = max(point["x_m"] for point in points)
        y0 = min(point["y_m"] for point in points)
        y1 = max(point["y_m"] for point in points)
        lines = []
        seen = set()

        def add_line(start, end, kind):
            key = tuple(sorted((
                (round(start[0], 6), round(start[1], 6)),
                (round(end[0], 6), round(end[1], 6)),
            )))
            if key in seen or ((end[0] - start[0]) ** 2 + (end[1] - start[1]) ** 2) <= 1e-10:
                return
            seen.add(key)
            lines.append({"kind": kind,
                          "start": {"x_m": key[0][0], "y_m": key[0][1]},
                          "end": {"x_m": key[1][0], "y_m": key[1][1]}})

        for polygon in polygons.values():
            for index, start in enumerate(polygon):
                end = polygon[(index + 1) % len(polygon)]
                if abs(start[0] - end[0]) > 1e-6 and abs(start[1] - end[1]) > 1e-6:
                    add_line(start, end, "outer_partition")

        rays = []
        for void in slab.get("voids", []):
            corners = [
                (void["x_min_m"], void["y_min_m"]), (void["x_max_m"], void["y_min_m"]),
                (void["x_max_m"], void["y_max_m"]), (void["x_min_m"], void["y_max_m"]),
            ]
            for x, y in corners:
                for dx, dy in ((-1, -1), (1, -1), (1, 1), (-1, 1)):
                    distances = []
                    if dx < 0:
                        distances.append((x0 - x) / dx)
                    else:
                        distances.append((x1 - x) / dx)
                    if dy < 0:
                        distances.append((y0 - y) / dy)
                    else:
                        distances.append((y1 - y) / dy)
                    distance = min(value for value in distances if value >= 0.0)
                    rays.append({"x": x, "y": y, "dx": dx, "dy": dy, "distance": distance})

        def cross(first, second):
            return first["dx"] * second["dy"] - first["dy"] * second["dx"]

        for index, first in enumerate(rays):
            for second in rays[index + 1:]:
                denominator = cross(first, second)
                if abs(denominator) < 1e-12:
                    continue
                rx, ry = second["x"] - first["x"], second["y"] - first["y"]
                t = (rx * second["dy"] - ry * second["dx"]) / denominator
                u = (rx * first["dy"] - ry * first["dx"]) / denominator
                if t > 1e-6 and t < first["distance"] - 1e-6 and u > 1e-6 and u < second["distance"] - 1e-6:
                    first["distance"] = min(first["distance"], t)
                    second["distance"] = min(second["distance"], u)

        for ray in rays:
            add_line((ray["x"], ray["y"]),
                     (ray["x"] + ray["dx"] * ray["distance"],
                      ray["y"] + ray["dy"] * ray["distance"]),
                     "void_partition")
        return lines

    def global_tributary_partition(slab):
        polygons = tributary_polygons(slab)
        gross_area = polygon_area_xy([
            (point["x_m"], point["y_m"]) for point in slab["coordinates"]
        ])
        partition_area = sum(polygon_area_xy(polygon) for polygon in polygons.values())
        overlap_area = sum(
            polygon_intersection_area(first, second)
            for index, first in enumerate(polygons.values())
            for second in list(polygons.values())[index + 1:]
        )
        return {
            "method": "global_45_degree_partition",
            "regions": [
                {"edge": edge, "polygon": [{"x_m": round(x, 6), "y_m": round(y, 6)} for x, y in polygon]}
                for edge, polygon in polygons.items()
            ],
            "area_m2": round(partition_area, 6),
            "area_check_m2": round(partition_area - gross_area, 6),
            "overlap_m2": round(overlap_area, 6),
        }

    def load_regions_for_slab(slab):
        x0 = min(point["x_m"] for point in slab["coordinates"])
        x1 = max(point["x_m"] for point in slab["coordinates"])
        y0 = min(point["y_m"] for point in slab["coordinates"])
        y1 = max(point["y_m"] for point in slab["coordinates"])
        x_breaks = {x0, x1}
        y_breaks = {y0, y1}
        level = next((item for item in loads.get("levels", [])
                      if abs(item["z_m"] - slab["z_m"]) < 1e-6), None)
        active_lt2 = (lt2_geometries if abs(slab["z_m"] - 19.8) < 1e-6
                      else lt2_geometries[:6])
        geometries = []
        if level:
            geometries.extend(geometry for zone in level.get("zones", [])
                              for geometry in zone.get("geometries", []))
        geometries.extend(active_lt2)
        for geometry in geometries:
            for point in geometry.get("points", []):
                if x0 < point[0] < x1:
                    x_breaks.add(point[0])
                if y0 < point[1] < y1:
                    y_breaks.add(point[1])

        def zone_at(point):
            if any(
                geometry.get("type") == "void"
                and point_in_polygon(point, geometry.get("points", []))
                for zone in (level or {}).get("zones", [])
                for geometry in zone.get("geometries", [])
            ) or any(
                geometry.get("type") == "void"
                and point_in_polygon(point, geometry.get("points", []))
                for geometry in active_lt2
            ):
                return None
            matches = [zone for zone in (level or {}).get("zones", [])
                       if zone_contains(zone, point)]
            if len(matches) == 1:
                return matches[0]
            inside_level_outer = any(
                geometry.get("type") == "polygon"
                and point_in_polygon(point, geometry.get("points", []))
                for zone in (level or {}).get("zones", [])
                for geometry in zone.get("geometries", [])
            )
            if inside_level_outer:
                return None
            return lt2_zone_for(point, slab["z_m"])

        def point_is_void(point):
            if any(
                geometry.get("type") == "void"
                and point_in_polygon(point, geometry.get("points", []))
                for zone in (level or {}).get("zones", [])
                for geometry in zone.get("geometries", [])
            ):
                return True
            return any(
                geometry.get("type") == "void"
                and point_in_polygon(point, geometry.get("points", []))
                for geometry in active_lt2
            )

        def nearest_zone(point):
            candidates = []
            for zone in (level or {}).get("zones", []):
                for geometry in zone.get("geometries", []):
                    if geometry.get("type") == "polygon":
                        candidates.append((zone, geometry["points"]))
            lt2_names = ((0, "LT2 A"), (1, "LT2 B"), (3, "LT2 C"), (4, "LT2 D"), (5, "LT2 E"))
            for index, name in lt2_names:
                if index >= len(active_lt2):
                    continue
                geometry = active_lt2[index]
                if geometry.get("type") == "polygon":
                    sc = 200.0 if name == "LT2 E" else (300.0 if name == "LT2 D" else 500.0)
                    candidates.append(({"id": name, "pm_adic_kg_m2": 260.0, "sc_kg_m2": sc}, geometry["points"]))

            def distance(bounds):
                bx0, bx1 = min(p[0] for p in bounds), max(p[0] for p in bounds)
                by0, by1 = min(p[1] for p in bounds), max(p[1] for p in bounds)
                dx = max(bx0 - point[0], 0.0, point[0] - bx1)
                dy = max(by0 - point[1], 0.0, point[1] - by1)
                return dx * dx + dy * dy

            return min(candidates, key=lambda item: distance(item[1]))[0] if candidates else None

        regions = {}
        partition_faces = []
        global_partition = global_tributary_partition(slab)
        edge_polygons = {
            region["edge"]: [(point["x_m"], point["y_m"]) for point in region["polygon"]]
            for region in global_partition["regions"]
        }
        slab["global_partition"] = global_partition
        slab["global_partition"]["construction_lines_45"] = construction_lines_45(slab, edge_polygons)
        slab["tributary_geometry"] = {}
        for edge, polygon in edge_polygons.items():
            intersections = []
            for void in slab.get("voids", []):
                clipped = polygon_rectangle_intersection_polygon(
                    polygon, void["x_min_m"], void["x_max_m"], void["y_min_m"], void["y_max_m"]
                )
                if polygon_area_xy(clipped) > 1e-9:
                    intersections.append([{"x_m": round(x, 6), "y_m": round(y, 6)} for x, y in clipped])
            slab["tributary_geometry"][edge] = {
                "gross_polygon": [{"x_m": round(x, 6), "y_m": round(y, 6)} for x, y in polygon],
                "void_intersections": intersections,
            }
        xs, ys = sorted(x_breaks), sorted(y_breaks)
        for xa, xb in zip(xs, xs[1:]):
            for ya, yb in zip(ys, ys[1:]):
                if xb - xa <= 1e-6 or yb - ya <= 1e-6:
                    continue
                point = ((xa + xb) / 2.0, (ya + yb) / 2.0)
                zone = zone_at(point)
                if not zone:
                    continue
                key = zone["id"]
                region = regions.setdefault(key, {"zone": zone, "area_m2": 0.0, "edge_areas": {}})
                cell_area = (xb - xa) * (yb - ya)
                region["area_m2"] += cell_area
                for edge, polygon in edge_polygons.items():
                    clipped = polygon_rectangle_intersection_polygon(polygon, xa, xb, ya, yb)
                    contribution = polygon_area_xy(clipped)
                    region["edge_areas"][edge] = region["edge_areas"].get(edge, 0.0) + contribution
                    if contribution > 1e-9:
                        partition_faces.append({
                            "edge": edge,
                            "zone": zone["id"],
                            "area_m2": round(contribution, 6),
                            "polygon": [{"x_m": round(x, 6), "y_m": round(y, 6)} for x, y in clipped],
                        })
        slab["global_partition"]["faces"] = partition_faces
        slab["global_partition"]["loaded_area_m2"] = round(sum(face["area_m2"] for face in partition_faces), 6)
        slab["global_partition"]["loaded_area_check_m2"] = round(
            slab["global_partition"]["loaded_area_m2"] - slab["area_m2"], 6
        )
        return list(regions.values())

    beam_load_cases = []
    unassigned_slabs = []
    for slab in slabs:
        centroid = (
            sum(point["x_m"] for point in slab["coordinates"]) / 4.0,
            sum(point["y_m"] for point in slab["coordinates"]) / 4.0,
        )
        regions = load_regions_for_slab(slab)
        edge_areas = {}
        for region in regions:
            for edge, area in region["edge_areas"].items():
                edge_areas[edge] = edge_areas.get(edge, 0.0) + area
        for tributary in slab["tributary_loads"]:
            edge_area = edge_areas.get(tributary["edge"], 0.0)
            original_area = tributary["tributary_area_m2"]
            scale = edge_area / original_area if original_area else 0.0
            tributary["tributary_area_m2"] = round(edge_area, 6)
            for key in ("w_start_kN_m", "w_max_kN_m", "w_end_kN_m", "total_load_kN"):
                tributary[key] = round(tributary[key] * scale, 6)
            slab["tributary_geometry"][tributary["edge"]]["loaded_area_m2"] = round(edge_area, 6)
        slab["area_check_m2"] = round(
            sum(load["tributary_area_m2"] for load in slab["tributary_loads"]) - slab["area_m2"], 6
        )
        slab["load_regions"] = [{"zone": region["zone"]["id"],
                                  "area_m2": round(region["area_m2"], 6)}
                                 for region in regions]
        if not regions:
            unassigned_slabs.append({"slab_id": slab["id"], "z_m": slab["z_m"], "centroid_xy_m": list(centroid),
                                     "reason": "no_unique_load_zone"})
            continue
        for region in regions:
            zone = region["zone"]
            q_pm = (zone.get("pm_adic_kg_m2") or 0.0) * KG_TO_KN
            q_sc = (zone.get("sc_kg_m2") or 0.0) * KG_TO_KN
            q_g = slab_self_weight + q_pm
            for tributary in slab["tributary_loads"]:
                edge_total = edge_areas.get(tributary["edge"], 0.0)
                region_edge_area = region["edge_areas"].get(tributary["edge"], 0.0)
                edge_share = region_edge_area / edge_total if edge_total else 0.0
                area = tributary["tributary_area_m2"] * edge_share
                for beam_id in tributary["beam_ids"]:
                    share = beam_share(tributary["beam_ids"], beam_id)
                    segment_area = area * share
                    beam_load_cases.append({
                        "slab_id": slab["id"], "beam_id": beam_id, "level_z_m": slab["z_m"],
                        "zone": zone["id"], "edge": tributary["edge"], "distribution": tributary["distribution"],
                        "tributary_area_m2": round(segment_area, 6), "q_G_kN_m2": round(q_g, 6), "q_SC_kN_m2": round(q_sc, 6),
                        "dead_load_kN": round(segment_area * q_g, 6), "live_load_kN": round(segment_area * q_sc, 6),
                        "w_G_start_kN_m": round(tributary["w_start_kN_m"] * edge_share * share * q_g / slab_self_weight, 6),
                        "w_G_max_kN_m": round(tributary["w_max_kN_m"] * edge_share * share * q_g / slab_self_weight, 6),
                        "w_G_end_kN_m": round(tributary["w_end_kN_m"] * edge_share * share * q_g / slab_self_weight, 6),
                        "w_SC_start_kN_m": round(tributary["w_start_kN_m"] * edge_share * share * q_sc / slab_self_weight, 6),
                        "w_SC_max_kN_m": round(tributary["w_max_kN_m"] * edge_share * share * q_sc / slab_self_weight, 6),
                        "w_SC_end_kN_m": round(tributary["w_end_kN_m"] * edge_share * share * q_sc / slab_self_weight, 6),
                    })

    # Explicit cantilevers are geometry and load-transfer regions, not finite
    # elements. Their interior edge is supported by the beam line at y=16.15 m.
    explicit_cantilever_cases = []

    def support_beams_for_region(region):
        x0, x1 = region["x_min_m"], region["x_max_m"]
        y = region.get("support_y_m", region["y_min_m"])
        result = []
        for element in elements:
            if not element["type"].startswith("BEAM"):
                continue
            a = next(node for node in nodes if node["id"] == element["i"])
            b = next(node for node in nodes if node["id"] == element["j"])
            if abs(a["z_m"] - region["z_m"]) > 1e-6 or abs(b["z_m"] - region["z_m"]) > 1e-6:
                continue
            if abs(a["y_m"] - b["y_m"]) > 1e-6 or abs(a["y_m"] - y) > 1e-3:
                continue
            overlap = max(0.0, min(x1, max(a["x_m"], b["x_m"]))
                          - max(x0, min(a["x_m"], b["x_m"])))
            if overlap > 1e-6:
                result.append((element["id"], overlap, abs(b["x_m"] - a["x_m"])))
        return result

    for region in config.get("cantilever_regions", []):
        x0, x1 = region["x_min_m"], region["x_max_m"]
        y0, y1 = region["y_min_m"], region["y_max_m"]
        z = region["z_m"]
        area = (x1 - x0) * (y1 - y0)
        if area <= 1e-6:
            continue
        corners = [(x0, y0), (x1, y0), (x1, y1), (x0, y1)]
        slab_node_ids = []
        for x, y in corners:
            key = (round(x, 3), round(y, 3), round(z, 3))
            if key not in node_by_xyz:
                node = {"id": next_slab_node, "level": levels.index(z), "axis": "SLAB",
                        "x_m": x, "y_m": y, "z_m": z, "restraint": False,
                        "status": "MANUAL_SLAB_NODE"}
                next_slab_node += 1
                nodes.append(node)
                node_by_xyz[key] = node
            slab_node_ids.append(node_by_xyz[key]["id"])

        support_beams = support_beams_for_region(region)
        support_ids = [item[0] for item in support_beams]
        slab_id = next_slab_element
        next_slab_element += 1
        slabs.append({
            "id": slab_id, "panel_id": slab_id, "node_ids": slab_node_ids,
            "coordinates": [{"x_m": x, "y_m": y, "z_m": z} for x, y in corners],
            "edge_beam_ids": {"support": support_ids},
            "dimensions_m": {"lx": round(x1 - x0, 6), "ly": round(y1 - y0, 6)},
            "area_m2": round(area, 6), "gross_area_m2": round(area, 6),
            "thickness_m": slab_thickness, "density_kg_m3": slab_density,
            "self_weight_kN_m2": round(slab_self_weight, 6),
            "tributary_loads": [], "voids": [], "area_check_m2": 0.0,
            "load_regions": [{"zone": region["id"], "area_m2": round(area, 6)}],
            "boundary_elements": support_ids, "z_m": z,
            "status": "EXPLICIT_CANTILEVER",
        })

        if not support_beams:
            unassigned_slabs.append({"slab_id": slab_id, "z_m": z,
                                     "centroid_xy_m": [(x0 + x1) / 2.0, (y0 + y1) / 2.0],
                                     "reason": "cantilever_without_support_beam"})
            continue

        # Any short gap between the region and the beam-line extents is
        # assigned to the beam with the largest overlap to conserve the area.
        region_length = x1 - x0
        shares = {beam_id: overlap / region_length for beam_id, overlap, _ in support_beams}
        assigned_share = sum(shares.values())
        largest_beam = max(support_beams, key=lambda item: item[1])[0]
        shares[largest_beam] += max(0.0, 1.0 - assigned_share)
        q_pm = region.get("pm_adic_kg_m2", 0.0) * KG_TO_KN
        q_sc = region.get("sc_kg_m2", 0.0) * KG_TO_KN
        q_g = slab_self_weight + q_pm
        for beam_id, _, beam_length in support_beams:
            beam_area = area * shares[beam_id]
            dead_load = beam_area * q_g
            live_load = beam_area * q_sc
            w_g = dead_load / beam_length if beam_length else 0.0
            w_sc = live_load / beam_length if beam_length else 0.0
            explicit_cantilever_cases.append({
                "slab_id": slab_id, "beam_id": beam_id, "level_z_m": z,
                "zone": region["id"], "edge": "support",
                "distribution": "cantilever_to_support",
                "tributary_area_m2": round(beam_area, 6),
                "q_G_kN_m2": round(q_g, 6), "q_SC_kN_m2": round(q_sc, 6),
                "dead_load_kN": round(dead_load, 6),
                "live_load_kN": round(live_load, 6),
                "w_G_start_kN_m": round(w_g, 6), "w_G_max_kN_m": round(w_g, 6),
                "w_G_end_kN_m": round(w_g, 6), "w_SC_start_kN_m": round(w_sc, 6),
                "w_SC_max_kN_m": round(w_sc, 6), "w_SC_end_kN_m": round(w_sc, 6),
            })

    beam_load_cases.extend(explicit_cantilever_cases)

    # Close the remaining loaded source cells that are not covered by a
    # generated slab. These panels are load-transfer geometry only; they do not
    # add finite elements. The nearest beam receives the complete cell load.
    def source_zone_at(point, z):
        level_zones = zones_by_level.get(z, [])
        if any(
            geometry.get("type") == "void"
            and point_in_polygon(point, geometry.get("points", []))
            for zone in level_zones
            for geometry in zone.get("geometries", [])
        ):
            return None
        matches = [zone for zone in level_zones if zone_contains(zone, point)]
        if len(matches) == 1:
            return matches[0]
        inside_level_polygon = any(
            geometry.get("type") == "polygon"
            and point_in_polygon(point, geometry.get("points", []))
            for zone in level_zones
            for geometry in zone.get("geometries", [])
        )
        return None if inside_level_polygon else lt2_zone_for(point, z)

    def slab_covers_point(slab, point):
        xs = [item["x_m"] for item in slab["coordinates"]]
        ys = [item["y_m"] for item in slab["coordinates"]]
        if not min(xs) <= point[0] <= max(xs) or not min(ys) <= point[1] <= max(ys):
            return False
        return not any(
            void["x_min_m"] < point[0] < void["x_max_m"]
            and void["y_min_m"] < point[1] < void["y_max_m"]
            for void in slab.get("voids", [])
        )

    def point_segment_distance(point, a, b):
        dx, dy = b["x_m"] - a["x_m"], b["y_m"] - a["y_m"]
        length_squared = dx * dx + dy * dy
        if length_squared <= 1e-12:
            return ((point[0] - a["x_m"]) ** 2 + (point[1] - a["y_m"]) ** 2) ** 0.5
        ratio = ((point[0] - a["x_m"]) * dx + (point[1] - a["y_m"]) * dy) / length_squared
        ratio = max(0.0, min(1.0, ratio))
        return ((point[0] - (a["x_m"] + ratio * dx)) ** 2
                + (point[1] - (a["y_m"] + ratio * dy)) ** 2) ** 0.5

    def nearest_beam(point, z):
        candidates = []
        for element in elements:
            if not element["type"].startswith("BEAM"):
                continue
            a = next(node for node in nodes if node["id"] == element["i"])
            b = next(node for node in nodes if node["id"] == element["j"])
            if abs(a["z_m"] - z) > 1e-6 or abs(b["z_m"] - z) > 1e-6:
                continue
            length = ((a["x_m"] - b["x_m"]) ** 2 + (a["y_m"] - b["y_m"]) ** 2) ** 0.5
            if length > 1e-6:
                candidates.append((point_segment_distance(point, a, b), element["id"], length))
        return min(candidates) if candidates else None

    missing_load_panels = 0
    missing_load_area = 0.0
    missing_load_cases = []
    for z in levels:
        level_zones = zones_by_level.get(z, [])
        active_geometries = [geometry for zone in level_zones
                             for geometry in zone.get("geometries", [])]
        if z in loads.get("lt2_levels", []) or abs(z - 19.8) < 1e-6:
            active_geometries.extend(
                lt2_geometries if abs(z - 19.8) < 1e-6 else lt2_geometries[:6]
            )
        x_breaks = {round(point["x_m"], 6) for slab in slabs if abs(slab["z_m"] - z) < 1e-6
                    for point in slab["coordinates"]}
        y_breaks = {round(point["y_m"], 6) for slab in slabs if abs(slab["z_m"] - z) < 1e-6
                    for point in slab["coordinates"]}
        x_breaks.update(round(point[0], 6) for geometry in active_geometries
                        for point in geometry.get("points", []))
        y_breaks.update(round(point[1], 6) for geometry in active_geometries
                        for point in geometry.get("points", []))
        for xa, xb in zip(sorted(x_breaks), sorted(x_breaks)[1:]):
            for ya, yb in zip(sorted(y_breaks), sorted(y_breaks)[1:]):
                if xb - xa <= 1e-6 or yb - ya <= 1e-6:
                    continue
                point = ((xa + xb) / 2.0, (ya + yb) / 2.0)
                zone = source_zone_at(point, z)
                if not zone or any(slab_covers_point(slab, point)
                                   for slab in slabs if abs(slab["z_m"] - z) < 1e-6):
                    continue
                area = (xb - xa) * (yb - ya)
                support = nearest_beam(point, z)
                if support is None:
                    unassigned_slabs.append({
                        "slab_id": None, "z_m": z, "centroid_xy_m": list(point),
                        "reason": "missing_source_area_without_support_beam",
                    })
                    continue
                beam_id, beam_length = support[1], support[2]
                node_ids = []
                for x, y in ((xa, ya), (xb, ya), (xb, yb), (xa, yb)):
                    key = (round(x, 3), round(y, 3), round(z, 3))
                    if key not in node_by_xyz:
                        node = {"id": next_slab_node, "level": levels.index(z), "axis": "SLAB",
                                "x_m": x, "y_m": y, "z_m": z, "restraint": False,
                                "status": "MISSING_LOAD_PANEL_NODE"}
                        next_slab_node += 1
                        nodes.append(node)
                        node_by_xyz[key] = node
                    node_ids.append(node_by_xyz[key]["id"])
                slab_id = next_slab_element
                next_slab_element += 1
                slabs.append({
                    "id": slab_id, "panel_id": slab_id, "node_ids": node_ids,
                    "coordinates": [{"x_m": x, "y_m": y, "z_m": z}
                                    for x, y in ((xa, ya), (xb, ya), (xb, yb), (xa, yb))],
                    "edge_beam_ids": {"nearest_support": [beam_id]},
                    "dimensions_m": {"lx": round(xb - xa, 6), "ly": round(yb - ya, 6)},
                    "area_m2": round(area, 6), "gross_area_m2": round(area, 6),
                    "thickness_m": slab_thickness, "density_kg_m3": slab_density,
                    "self_weight_kN_m2": round(slab_self_weight, 6),
                    "tributary_loads": [], "voids": [], "area_check_m2": 0.0,
                    "load_regions": [{"zone": zone["id"], "area_m2": round(area, 6)}],
                    "boundary_elements": [beam_id], "z_m": z,
                    "status": "EXPLICIT_LOAD_PANEL",
                })
                q_g = slab_self_weight + zone.get("pm_adic_kg_m2", 0.0) * KG_TO_KN
                q_sc = zone.get("sc_kg_m2", 0.0) * KG_TO_KN
                dead_load = area * q_g
                live_load = area * q_sc
                missing_load_cases.append({
                    "slab_id": slab_id, "beam_id": beam_id, "level_z_m": z,
                    "zone": zone["id"], "edge": "nearest_support",
                    "distribution": "missing_area_to_nearest_beam",
                    "tributary_area_m2": round(area, 6), "q_G_kN_m2": round(q_g, 6),
                    "q_SC_kN_m2": round(q_sc, 6), "dead_load_kN": round(dead_load, 6),
                    "live_load_kN": round(live_load, 6),
                    "w_G_start_kN_m": round(dead_load / beam_length, 6),
                    "w_G_max_kN_m": round(dead_load / beam_length, 6),
                    "w_G_end_kN_m": round(dead_load / beam_length, 6),
                    "w_SC_start_kN_m": round(live_load / beam_length, 6),
                    "w_SC_max_kN_m": round(live_load / beam_length, 6),
                    "w_SC_end_kN_m": round(live_load / beam_length, 6),
                })
                missing_load_panels += 1
                missing_load_area += area
    beam_load_cases.extend(missing_load_cases)

    beam_slab_loads = []
    for slab in slabs:
        for load in slab["tributary_loads"]:
            for beam_id in load["beam_ids"]:
                share = beam_share(load["beam_ids"], beam_id)
                beam_slab_loads.append({
                    "slab_id": slab["id"], "beam_id": beam_id, "z_m": slab["z_m"],
                    "edge": load["edge"], "tributary_area_m2": round(load["tributary_area_m2"] * share, 6),
                    "distribution": load["distribution"],
                    "w_start_kN_m": round(load["w_start_kN_m"] * share, 6),
                    "w_max_kN_m": round(load["w_max_kN_m"] * share, 6),
                    "w_end_kN_m": round(load["w_end_kN_m"] * share, 6),
                    "total_load_kN": round(load["total_load_kN"] * share, 6),
                })

    data = {"source": "geometria_manual.json", "load_source": "cargas_losas.json", "status": "MANUAL_REVIEW",
             "units": "kN-m-s", "nodes": nodes, "elements": elements,
             "walls": config.get("walls", []),
             "slabs": slabs,
             "beam_slab_loads": beam_slab_loads,
             "load_zones": loads.get("levels", []),
             "lt2_load_zones": loads.get("lt2", {}),
             "lt2_load_cases": loads.get("lt2_load_cases", []),
             "load_zone_totals": load_zone_totals,
              "beam_load_cases": beam_load_cases,
              "unassigned_load_slabs": unassigned_slabs,
              "void_slabs": [{"slab_id": slab["id"], "z_m": slab["z_m"],
                              "coordinates": slab["coordinates"]} for slab in void_slabs],
            "section_columns": {"A_m2": 0.49, "Iy_m4": 0.020004, "Iz_m4": 0.020004, "J_m4": 0.040008},
            "section_beams": {"A_m2": 0.48, "Iy_m4": 0.0256, "Iz_m4": 0.0144, "J_m4": 0.002},
            "section_small_beams": {"A_m2": 0.135, "Iy_m4": 0.002278, "Iz_m4": 0.000506, "J_m4": 0.0002},
            "section_variable_beams": {"A_m2": 0.21, "Iy_m4": 0.005, "Iz_m4": 0.0015, "J_m4": 0.0004},
            "section_40x60_beams": {"A_m2": 0.24, "Iy_m4": 0.0072, "Iz_m4": 0.0032, "J_m4": 0.0008},
            "material": config["material"], "mass_per_node_t": config["mass_per_node_t"]}
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    MODEL.write_text(json.dumps(data, indent=2), encoding="utf-8")
    CSV.parent.mkdir(parents=True, exist_ok=True)
    lines = ["kind,id,type,i,j,x_m,y_m,z_m,level,status"]
    lines.extend(f'N,{n["id"]},NODE,,,,{n["x_m"]},{n["y_m"]},{n["z_m"]},{n["level"]},{n["status"]}' for n in nodes)
    lines.extend(f'E,{e["id"]},{e["type"]},{e["i"]},{e["j"]},,,,,,{e["status"]}' for e in elements)
    for wall in config.get("walls", []):
        lines.append(
            f'W,{wall["id"]},WALL,{wall["x_i_m"]},{wall["y_i_m"]},{wall["z_i_m"]},'
            f'{wall["x_j_m"]},{wall["y_j_m"]},{wall["z_j_m"]},{wall["thickness_m"]},{wall.get("status", "MANUAL")}'
        )
    for slab in slabs:
        n1, n2, n3, n4 = slab["node_ids"]
        lines.append(f'S,{slab["id"]},{n1},{n2},{n3},{n4},{slab["z_m"]},{slab["thickness_m"]},{slab["status"]}')
        for void in slab.get("voids", []):
            lines.append(
                f'V,{slab["id"]},{void["x_min_m"]},{void["x_max_m"]},'
                f'{void["y_min_m"]},{void["y_max_m"]}'
            )
    CSV.write_text("\n".join(lines) + "\n", encoding="utf-8")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nodos_Manual"
    sheet.append(["id", "nivel", "eje", "x_m", "y_m", "z_m", "restriccion", "estado"])
    for n in nodes:
        sheet.append([n["id"], n["level"], n["axis"], n["x_m"], n["y_m"], n["z_m"], "Empotrado" if n["restraint"] else "Libre", n["status"]])
    el_sheet = workbook.create_sheet("Elementos")
    el_sheet.append(["id", "tipo", "nodo_i", "nodo_j", "estado"])
    for e in elements:
        el_sheet.append([e["id"], e["type"], e["i"], e["j"], e["status"]])
    slab_sheet = workbook.create_sheet("Losas")
    slab_sheet.append(["id", "nodo_1", "nodo_2", "nodo_3", "nodo_4", "z_m", "lx_m", "ly_m", "area_m2", "espesor_m", "densidad_kg_m3", "q_kN_m2", "control_area_m2", "estado"])
    for slab in slabs:
        slab_sheet.append([slab["id"], *slab["node_ids"], slab["z_m"],
                           slab["dimensions_m"]["lx"], slab["dimensions_m"]["ly"], slab["area_m2"],
                           slab["thickness_m"], slab["density_kg_m3"], slab["self_weight_kN_m2"],
                           slab["area_check_m2"], slab["status"]])
    load_sheet = workbook.create_sheet("Cargas_Losa")
    load_sheet.append(["losa_id", "viga_id", "z_m", "borde", "area_tributaria_m2", "tipo", "w_inicio_kN_m", "w_max_kN_m", "w_final_kN_m", "carga_total_kN"])
    for load in beam_slab_loads:
        load_sheet.append([load["slab_id"], load["beam_id"], load["z_m"], load["edge"],
                           load["tributary_area_m2"], load["distribution"],
                           load["w_start_kN_m"], load["w_max_kN_m"], load["w_end_kN_m"],
                           load["total_load_kN"]])
    notes = workbook.create_sheet("Supuestos")
    notes.append(["campo", "valor"])
    notes.append(["niveles", "0, 3.96, 7.92, 11.88, 15.84, 19.80 m"])
    notes.append(["ejes Y", "Eje 3 = 0; Eje 2 = 7.25; Eje 1 = 16.15 m"])
    notes.append(["columnas", "Todas 0.70 x 0.70 m; datos ingresados manualmente"])
    notes.append(["vigas", "Cielo 1 subterraneo: 7; Cielos Piso 1 y Piso 2: base repetida; Piso 2 agrega voladizos; Piso 3 conserva la base y agrega su configuracion especial"])
    notes.append(["modelo", "Vigas, columnas, muros y losas cerradas por vigas; sin fundaciones"])
    workbook.save(EXCEL)
    print(f"Modelo manual: {len(nodes)} nodos, {len(elements)} elementos; Excel: {EXCEL.name}")


if __name__ == "__main__":
    main()
