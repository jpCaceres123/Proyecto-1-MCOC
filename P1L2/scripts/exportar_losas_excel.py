"""Exporta un Excel independiente para revisar e identificar losas."""

from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MODEL = ROOT / "outputs" / "modelo_3d_manual.json"
OUTPUT = ROOT / "outputs" / "losas_modelo.xlsx"


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 10), 45)


def format_vertices(slab):
    return " | ".join(
        f"({p['x_m']:.3f}, {p['y_m']:.3f}, {p['z_m']:.3f})"
        for p in slab["coordinates"]
    )


def main():
    data = json.loads(MODEL.read_text(encoding="utf-8"))
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Resumen_Losas"
    summary.append([
        "losa_id", "nivel_z_m", "nodos", "vertices_xyz_m", "lx_m", "ly_m",
        "area_m2", "espesor_m", "densidad_kg_m3", "q_kN_m2", "vigas_borde",
        "control_area_m2", "estado",
    ])
    for slab in data.get("slabs", []):
        edge_beams = "; ".join(
            f"{edge}: {','.join(str(beam_id) for beam_id in beam_ids)}"
            for edge, beam_ids in slab["edge_beam_ids"].items()
        )
        summary.append([
            slab["id"], slab["z_m"], ", ".join(str(node_id) for node_id in slab["node_ids"]),
            format_vertices(slab), slab["dimensions_m"]["lx"], slab["dimensions_m"]["ly"],
            slab["area_m2"], slab["thickness_m"], slab["density_kg_m3"],
            slab["self_weight_kN_m2"], edge_beams, slab["area_check_m2"], slab["status"],
        ])
    style_sheet(summary)

    loads = workbook.create_sheet("Cargas_Tributarias")
    loads.append([
        "losa_id", "nivel_z_m", "viga_id", "borde", "area_tributaria_m2",
        "tipo_carga", "w_inicio_kN_m", "w_max_kN_m", "w_final_kN_m", "carga_total_kN",
    ])
    for load in data.get("beam_slab_loads", []):
        loads.append([
            load["slab_id"], load["z_m"], load["beam_id"], load["edge"],
            load["tributary_area_m2"], load["distribution"], load["w_start_kN_m"],
            load["w_max_kN_m"], load["w_end_kN_m"], load["total_load_kN"],
        ])
    style_sheet(loads)

    levels = workbook.create_sheet("Resumen_Niveles")
    levels.append(["nivel_z_m", "cantidad_losas", "area_total_m2", "carga_total_kN"])
    by_level = {}
    for slab in data.get("slabs", []):
        item = by_level.setdefault(slab["z_m"], {"count": 0, "area": 0.0})
        item["count"] += 1
        item["area"] += slab["area_m2"]
    for z, item in sorted(by_level.items()):
        load = sum(
            slab["self_weight_kN_m2"] * slab["area_m2"]
            for slab in data.get("slabs", [])
            if slab["z_m"] == z
        )
        levels.append([z, item["count"], round(item["area"], 6), round(load, 6)])
    style_sheet(levels)

    workbook.save(OUTPUT)
    print(f"Excel de losas creado: {OUTPUT.name}")


if __name__ == "__main__":
    main()
