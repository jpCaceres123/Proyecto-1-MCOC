"""Genera nodos, elementos, Excel y CSV desde geometria_manual.json."""

from pathlib import Path
import json

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent
MODEL = ROOT / "modelo_3d_manual.json"
EXCEL = ROOT / "nodos_modelo_manual.xlsx"
CSV = ROOT / "UnityVisualization" / "Assets" / "Resources" / "model_3d.csv"


def main():
    config = json.loads((ROOT / "geometria_manual.json").read_text(encoding="utf-8"))
    nodes = []
    by_level = {z: [] for z in config["levels_m"]}
    for k, z in enumerate(config["levels_m"]):
        for y_index, y in enumerate(config["y_axes_m"]):
            for axis in config["column_axes"]:
                if z < axis["z_inicio_m"]:
                    continue
                tag = 100000 * k + 1000 * y_index + int(axis["x_m"] * 10) + 1
                node = {"id": tag, "level": k, "axis": axis["eje"],
                        "x_m": axis["x_m"], "y_m": y, "z_m": z,
                        "restraint": k == 0, "status": "MANUAL"}
                nodes.append(node)
                by_level[z].append(node)

        # Some columns occupy only selected intersections of the extended grid.
        # They are declared explicitly instead of being created at every Y row.
        for point_index, point in enumerate(config.get("column_points", [])):
            if z < point.get("z_inicio_m", 0.0):
                continue
            tag = 700000 + 1000 * k + point_index + 1
            node = {"id": tag, "level": k, "axis": point["id"],
                    "x_m": point["x_m"], "y_m": point["y_m"], "z_m": z,
                    "restraint": k == 0, "status": "MANUAL"}
            nodes.append(node)
            by_level[z].append(node)

    manual_beams = list(config.get("manual_beams", []))
    for repeat in config.get("repeat_manual_beams", []):
        def allowed(beam):
            if not repeat.get("exclude_negative_y", False):
                return True
            return all(beam.get(key, 0.0) >= 0.0 for key in ("y_m", "y_i_m", "y_j_m") if key in beam)

        manual_beams.extend(
            [dict(beam, z_m=repeat["to_z_m"])
             for beam in manual_beams
             if round(beam["z_m"], 3) == round(repeat["from_z_m"], 3)
             and (not repeat.get("source_group") or beam.get("group") == repeat["source_group"])
             and beam.get("group") != repeat.get("exclude_group")
             and allowed(beam)]
        )

    elements = []
    eid = 1
    levels = config["levels_m"]
    for z0, z1 in zip(levels, levels[1:]):
        low = {(n["axis"], round(n["y_m"], 3)): n for n in by_level[z0]}
        high = {(n["axis"], round(n["y_m"], 3)): n for n in by_level[z1]}
        for key, top in high.items():
            if key in low:
                elements.append({"id": eid, "type": "COLUMN", "i": low[key]["id"], "j": top["id"], "status": "MANUAL"})
                eid += 1

    manual_z = {round(b["z_m"], 3) for b in manual_beams}
    node_at = {(n["axis"], round(n["y_m"], 3), round(n["z_m"], 3)): n for n in nodes}
    node_by_xyz = {(round(n["x_m"], 3), round(n["y_m"], 3), round(n["z_m"], 3)): n for n in nodes}
    next_manual_node = 900000

    def ensure_beam_node(x, y, z):
        nonlocal next_manual_node
        key = (round(x, 3), round(y, 3), round(z, 3))
        if key not in node_by_xyz:
            node = {"id": next_manual_node, "level": levels.index(z), "axis": "BEAM_MANUAL",
                    "x_m": x, "y_m": y, "z_m": z, "restraint": False, "status": "MANUAL_BEAM_NODE"}
            next_manual_node += 1
            nodes.append(node)
            node_by_xyz[key] = node
        return node_by_xyz[key]

    for beam in manual_beams:
        z = round(beam["z_m"], 3)
        if "x_i_m" in beam:
            a = ensure_beam_node(beam["x_i_m"], beam["y_i_m"], z)
            b = ensure_beam_node(beam["x_j_m"], beam["y_j_m"], z)
        elif beam["direction"] == "X":
            a = node_at[(beam["axis_i"], round(beam["y_m"], 3), z)]
            b = node_at[(beam["axis_j"], round(beam["y_m"], 3), z)]
        else:
            x = next(axis["x_m"] for axis in config["column_axes"] if axis["eje"] == beam["axis"])
            a = node_at.get((beam["axis"], round(beam["y_i_m"], 3), z),
                            ensure_beam_node(x, beam["y_i_m"], z))
            b = node_at.get((beam["axis"], round(beam["y_j_m"], 3), z),
                            ensure_beam_node(x, beam["y_j_m"], z))
        elements.append({"id": eid, "type": beam.get("member_type", "BEAM_" + beam["direction"]), "direction": beam["direction"], "i": a["id"], "j": b["id"], "status": "MANUAL"})
        eid += 1

    # Provisional beams: connect adjacent listed column axes and Y rows where no manual layout exists.
    for z in levels[1:]:
        if round(z, 3) in manual_z:
            continue
        current = by_level[z]
        for y in config["y_axes_m"]:
            row = sorted((n for n in current if abs(n["y_m"] - y) < 1e-9), key=lambda n: n["x_m"])
            for a, b in zip(row, row[1:]):
                elements.append({"id": eid, "type": "BEAM_X", "direction": "X", "i": a["id"], "j": b["id"], "status": "PROVISIONAL_BEAM"})
                eid += 1
        for axis in config["column_axes"]:
            col = sorted((n for n in current if n["axis"] == axis["eje"]), key=lambda n: n["y_m"])
            for a, b in zip(col, col[1:]):
                elements.append({"id": eid, "type": "BEAM_Y", "direction": "Y", "i": a["id"], "j": b["id"], "status": "PROVISIONAL_BEAM"})
                eid += 1

    data = {"source": "geometria_manual.json", "status": "MANUAL_REVIEW",
            "units": "kN-m-s", "nodes": nodes, "elements": elements,
            "walls": config.get("walls", []),
            "section_columns": {"A_m2": 0.49, "Iy_m4": 0.020004, "Iz_m4": 0.020004, "J_m4": 0.040008},
            "section_beams": {"A_m2": 0.48, "Iy_m4": 0.0256, "Iz_m4": 0.0144, "J_m4": 0.002},
            "section_small_beams": {"A_m2": 0.135, "Iy_m4": 0.002278, "Iz_m4": 0.000506, "J_m4": 0.0002},
            "section_variable_beams": {"A_m2": 0.21, "Iy_m4": 0.005, "Iz_m4": 0.0015, "J_m4": 0.0004},
            "section_40x60_beams": {"A_m2": 0.24, "Iy_m4": 0.0072, "Iz_m4": 0.0032, "J_m4": 0.0008},
            "material": config["material"], "mass_per_node_t": config["mass_per_node_t"]}
    MODEL.write_text(json.dumps(data, indent=2), encoding="utf-8")
    CSV.parent.mkdir(parents=True, exist_ok=True)
    lines = ["kind,id,type,i,j,x_m,y_m,z_m,level,status"]
    lines.extend(f'N,{n["id"]},NODE,,,,{n["x_m"]},{n["y_m"]},{n["z_m"]},{n["level"]},{n["status"]}' for n in nodes)
    lines.extend(f'E,{e["id"]},{e["type"]},{e["i"]},{e["j"]},,,,,,{e["status"]}' for e in elements)
    for wall in config.get("walls", []):
        lines.append(
            f'W,{wall["id"]},WALL,{wall["x_i_m"]},{wall["y_i_m"]},{wall["z_i_m"]},'
            f'{wall["x_j_m"]},{wall["y_j_m"]},{wall["z_j_m"]},{wall["thickness_m"]},{wall.get("status", "MANUAL")}'
        )
    CSV.write_text("\n".join(lines) + "\n", encoding="utf-8")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nodos_Manual"
    sheet.append(["id", "nivel", "eje", "x_m", "y_m", "z_m", "restriccion", "estado"])
    for n in nodes:
        sheet.append([n["id"], n["level"], n["axis"], n["x_m"], n["y_m"], n["z_m"], "Empotrado" if n["restraint"] else "Libre", n["status"]])
    el_sheet = workbook.create_sheet("Elementos")
    el_sheet.append(["id", "tipo", "nodo_i", "nodo_j", "estado"])
    for e in elements:
        el_sheet.append([e["id"], e["type"], e["i"], e["j"], e["status"]])
    notes = workbook.create_sheet("Supuestos")
    notes.append(["campo", "valor"])
    notes.append(["niveles", "0, 3.96, 7.92, 11.88, 15.84, 19.80 m"])
    notes.append(["ejes Y", "Eje 3 = 0; Eje 2 = 7.25; Eje 1 = 16.15 m"])
    notes.append(["columnas", "Todas 0.70 x 0.70 m; datos ingresados manualmente"])
    notes.append(["vigas", "Cielo 1 subterraneo: 7; Cielos Piso 1 y Piso 2: base repetida; Piso 2 agrega voladizos; Piso 3 conserva la base y agrega su configuracion especial"])
    notes.append(["modelo", "Solo vigas y columnas; sin losas, muros ni fundaciones"])
    workbook.save(EXCEL)
    print(f"Modelo manual: {len(nodes)} nodos, {len(elements)} elementos; Excel: {EXCEL.name}")


if __name__ == "__main__":
    main()
